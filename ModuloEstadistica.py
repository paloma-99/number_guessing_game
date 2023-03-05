import openpyxl
import matplotlib.pyplot as plt

def Resultados(intentos_restantes, intentos_totales):
    file = openpyxl.load_workbook("estadisticas.xlsx")
    
    nombre_jugador = str(input("\nNombre: "))
    
    print("\nPodrá ver sus resultados en la sección de Estadísticas ingresando su nombre.")
    print("\n*********************************************************************************************")

    intentos_usados = intentos_totales - 1 - intentos_restantes

    if intentos_restantes == 0:
        resultado = "Perdedor"
    else:
        resultado = "Ganador"

    if intentos_totales == 21:
        dificultad = "Facil"
    elif intentos_totales == 13:
        dificultad = "Media"
    else:
        dificultad = "Dificil"
    
    hoja = file["Resultados"]
    hoja.append([nombre_jugador, resultado, intentos_usados, dificultad])
    file.save("estadisticas.xlsx")


def GanadosPerdidos():
    file = openpyxl.load_workbook("estadisticas.xlsx")
    hoja = file["Resultados"]

    nombre = str(input("\nNombre del jugador: "))

    ganador_facil = 0
    perdedor_facil = 0
    for i in range(1,hoja.max_row+1):
        if hoja.cell(row = i, column = 1).value == nombre:
            if hoja.cell(row = i, column = 4).value == "Facil":
                if hoja.cell(row = i, column = 2).value == "Ganador":
                    ganador_facil = ganador_facil + 1
                else:
                    perdedor_facil = perdedor_facil + 1

    ganador_media = 0
    perdedor_media = 0
    for i in range(1,hoja.max_row+1):
        if hoja.cell(row = i, column = 1).value == nombre:
            if hoja.cell(row = i, column = 4).value == "Media":
                if hoja.cell(row = i, column = 2).value == "Ganador":
                    ganador_media = ganador_media + 1
                else:
                    perdedor_media = perdedor_media + 1

    ganador_dificil = 0
    perdedor_dificil = 0
    for i in range(1,hoja.max_row+1):
        if hoja.cell(row = i, column = 1).value == nombre:
            if hoja.cell(row = i, column = 4).value == "Dificil":
                if hoja.cell(row = i, column = 2).value == "Ganador":
                    ganador_dificil = ganador_dificil + 1
                else:
                    perdedor_dificil = perdedor_dificil + 1
    
    x = ["Ganados" , "Perdidos"]
    y_facil = [ganador_facil , perdedor_facil]
    y_media = [ganador_media , perdedor_media]
    y_dificil = [ganador_dificil , perdedor_dificil]

    fig, (ax1, ax2, ax3) = plt.subplots(1, 3)
    fig.suptitle("Jugador: " + nombre)
    ax1.bar(x, y_facil)
    ax1.set_title("Fácil")
    ax2.bar(x, y_media)
    ax2.set_title("Media")
    ax3.bar(x, y_dificil)
    ax3.set_title("Difícil")
    fig.tight_layout()
    plt.show()


def Intentos():
    file = openpyxl.load_workbook("estadisticas.xlsx")
    hoja = file["Resultados"]

    nombre = str(input("\nNombre del jugador: "))

    intentos_facil = []
    for i in range(1,hoja.max_row+1):
        if hoja.cell(row = i, column = 1).value == nombre:
            if hoja.cell(row = i, column = 4).value == "Facil":
                if hoja.cell(row = i, column = 2).value == "Ganador":
                    intentos_facil.append(hoja.cell(row = i, column = 3).value)

    intentos_media = []
    for i in range(1,hoja.max_row+1):
        if hoja.cell(row = i, column = 1).value == nombre:
            if hoja.cell(row = i, column = 4).value == "Media":
                if hoja.cell(row = i, column = 2).value == "Ganador":
                    intentos_media.append(hoja.cell(row = i, column = 3).value)

    intentos_dificil = []
    for i in range(1,hoja.max_row+1):
        if hoja.cell(row = i, column = 1).value == nombre:
            if hoja.cell(row = i, column = 4).value == "Dificil":
                if hoja.cell(row = i, column = 2).value == "Ganador":
                    intentos_dificil.append(hoja.cell(row = i, column = 3).value)

    fig, (ax1, ax2, ax3) = plt.subplots(1, 3)
    fig.suptitle("Jugador: " + nombre)
    ax1.hist(intentos_facil)
    ax1.set_title("Fácil")
    ax2.hist(intentos_media)
    ax2.set_title("Media")
    ax3.hist(intentos_dificil)
    ax3.set_title("Difícil")
    fig.tight_layout()
    plt.show()